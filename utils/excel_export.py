import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_gap_excel(result: dict, policy_name: str) -> bytes:
    """Generates a styled Excel file for Gap Analysis."""
    wb = Workbook()
    BLUE = "1E3A5F"
    LIGHT_BLUE = "DCE6F1"
    RED = "FCE4D6"
    GREEN = "E2EFDA"
    AMBER = "FFF2CC"
    
    ws = wb.active
    ws.title = "Gap Analysis"
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = "COMPLY.AI — Gap Analysis Report"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", fgColor=BLUE)
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30
    
    # Score
    score = result.get('compliance_score', 0)
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Compliance Score: {score}%  |  Policy: {policy_name}"
    ws['A2'].font = Font(size=12)
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Headers
    headers = ["#", "Obligation", "Severity", "What is Missing", "Rationale", "Recommendation"]
    for col, head in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=head)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2D6BE4")
        cell.alignment = Alignment(horizontal="center")
    
    # Gaps
    gaps = result.get('gaps', [])
    for i, gap in enumerate(gaps):
        row = 5 + i
        sev = gap.get('severity', 'medium').lower()
        fill_color = RED if sev == 'high' else AMBER if sev == 'medium' else GREEN
        ws.cell(row=row, column=1, value=i+1)
        ws.cell(row=row, column=2, value=gap.get('obligation', ''))
        ws.cell(row=row, column=3, value=gap.get('severity', '').upper())
        ws.cell(row=row, column=4, value=gap.get('what_is_missing', ''))
        ws.cell(row=row, column=5, value=gap.get('rationale', ''))
        ws.cell(row=row, column=6, value=gap.get('recommended_action', ''))
        
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
        ws.row_dimensions[row].height = 60
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 35
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def build_policy_excel(guidance: dict, company_name: str) -> bytes:
    """Generates a styled Excel file for Policy Guidance."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Policy Guidance"
    
    ws.merge_cells('A1:E1')
    ws['A1'] = "COMPLY.AI — Custom Policy Guidance"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", fgColor="1E3A5F")
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:E2')
    ws['A2'] = f"Company: {company_name}  |  Readiness Score: {guidance.get('readiness_score', 0)}%"
    ws['A2'].alignment = Alignment(horizontal="center")
    
    headers = ["Section", "What to Include", "Sample Clause", "Why It Matters", "Reg. Ref."]
    for col, head in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=head)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2D6BE4")
        cell.alignment = Alignment(horizontal="center")
    
    sections = guidance.get('sections', [])
    for i, sec in enumerate(sections):
        row = 5 + i
        ws.cell(row=row, column=1, value=sec.get('section_name', ''))
        ws.cell(row=row, column=2, value=sec.get('what_to_include', ''))
        ws.cell(row=row, column=3, value=sec.get('sample_clause', ''))
        ws.cell(row=row, column=4, value=sec.get('why_it_matters', ''))
        ws.cell(row=row, column=5, value=sec.get('regulation_reference', ''))
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
        ws.row_dimensions[row].height = 80
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 20
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()